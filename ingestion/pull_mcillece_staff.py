"""Parse a McIllece CFB Coaches Database file (CSV or XLSX).

Accepts a file path and returns a cleaned list of staff record dicts
with fields normalised for downstream Neo4j loading.

Usage (standalone):
    python -m ingestion.pull_mcillece_staff path/to/file.xlsx
"""

import logging
import sys
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_REQUIRED_FIELDS = {"coach_code", "team_code", "year", "team", "coach"}
_ROLE_COLS = ["pos1", "pos2", "pos3", "pos4", "pos5"]
_NULL_STRINGS = {"", "none", "nan", "n/a"}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def load_mcillece_file(file_path: str | Path) -> list[dict[str, Any]]:
    """Read a McIllece XLSX or CSV file and return cleaned staff records.

    Each returned dict has the keys:
    - ``coach_code`` (int) — unique coach identifier from the dataset
    - ``team_code`` (int) — unique team identifier from the dataset
    - ``year`` (int) — season year
    - ``team`` (str) — school name (canonical McIllece spelling)
    - ``coach_name`` (str) — full name as spelled in the dataset
    - ``roles`` (list[str]) — non-null values from pos1–pos5 columns

    Malformed rows (missing required fields or non-numeric codes) are
    skipped with a ``WARNING`` log and excluded from the output.

    Args:
        file_path: Path to a ``.csv`` or ``.xlsx`` McIllece data file.

    Returns:
        Cleaned list of staff record dicts, one per valid input row.

    Raises:
        ValueError: If the file extension is not ``.csv`` or ``.xlsx``.
        ImportError: If ``openpyxl`` is not installed and an XLSX file is given.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        raw_rows = _read_xlsx(path)
    elif suffix == ".csv":
        raw_rows = _read_csv(path)
    else:
        raise ValueError(
            f"Unsupported file extension '{suffix}'. Expected '.xlsx' or '.csv'."
        )

    return _clean_rows(raw_rows)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    """Read an XLSX file and return a list of row dicts with lower-cased keys."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read .xlsx files. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [
        str(cell.value).strip().lower() if cell.value is not None else f"_col{i}"
        for i, cell in enumerate(ws[1])
    ]

    rows: list[dict[str, Any]] = []
    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, raw_row)))

    return rows


def _read_csv(path: Path) -> list[dict[str, Any]]:
    """Read a CSV file and return a list of row dicts with lower-cased keys."""
    import csv

    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        return [
            {k.strip().lower(): v for k, v in row.items()}
            for row in reader
        ]


def _clean_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Validate and normalise raw rows from the file reader.

    Args:
        raw_rows: Dicts as produced by ``_read_xlsx`` or ``_read_csv``.

    Returns:
        Cleaned, validated list of staff record dicts.
    """
    cleaned: list[dict[str, Any]] = []

    for idx, row in enumerate(raw_rows, start=2):  # row 1 is header
        # Skip entirely blank rows
        if all(v is None or str(v).strip() == "" for v in row.values()):
            continue

        # Validate required fields are present and non-empty
        missing = [
            f for f in _REQUIRED_FIELDS
            if row.get(f) is None or str(row.get(f, "")).strip() == ""
        ]
        if missing:
            logger.warning("Row %d: skipping — missing required fields: %s", idx, missing)
            continue

        # Parse and validate numeric IDs
        try:
            coach_code = int(float(str(row["coach_code"])))
            team_code = int(float(str(row["team_code"])))
            year = int(float(str(row["year"])))
        except (ValueError, TypeError):
            logger.warning(
                "Row %d: skipping — non-numeric coach_code/team_code/year: "
                "coach_code=%r team_code=%r year=%r",
                idx,
                row.get("coach_code"),
                row.get("team_code"),
                row.get("year"),
            )
            continue

        # Extract roles from pos1–pos5, dropping nulls and sentinel strings
        roles = [
            str(row[col]).strip()
            for col in _ROLE_COLS
            if col in row
            and row[col] is not None
            and str(row[col]).strip().lower() not in _NULL_STRINGS
        ]

        cleaned.append(
            {
                "coach_code": coach_code,
                "team_code": team_code,
                "year": year,
                "team": str(row["team"]).strip(),
                "coach_name": str(row["coach"]).strip(),
                "roles": roles,
            }
        )

    logger.info(
        "Parsed %d valid staff records from %d raw rows", len(cleaned), len(raw_rows)
    )
    return cleaned


# ---------------------------------------------------------------------------
# Standalone entry-point
# ---------------------------------------------------------------------------


def _run(file_path: str) -> None:
    records = load_mcillece_file(file_path)
    print(f"Loaded {len(records):,} staff records")
    for r in records:
        print(r)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    if len(sys.argv) < 2:
        print("Usage: python -m ingestion.pull_mcillece_staff <file_path>")
        sys.exit(1)
    _run(sys.argv[1])
